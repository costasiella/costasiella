import io

from django.db.models import Q
from django.http import Http404, FileResponse
from django.utils.translation import gettext as _
import openpyxl

from ...models import Account
from ...modules.graphql_jwt_tools import get_user_from_cookie
from ...modules.gql_tools import get_rid


def _export_excel_accounts_active_get_header_mailinglist():
    return [
        _('First name'),
        _('Last name'),
        _('Email'),
    ]


def _export_excel_accounts_active_get_header_info():
    return [
        _('First name'),
        _('Last name'),
        _('Email'),
        _('Phone'),
        _('Mobile'),
        _('Emergency'),
        _('Date of birth'),
        _('Address'),
        _('Postcode'),
        _('City'),
        _('Country'),
        _('Door keynr')
    ]


def export_excel_accounts_active(request,**kwargs):
    """
    Export active accounts
    """
    user = get_user_from_cookie(request)
    if not user.has_perm('costasiella.view_account'):
        raise Http404("Permission denied")

    wb = openpyxl.Workbook(write_only=True)
    ws_info = wb.create_sheet(_("Active accounts"))
    ws_info.append(_export_excel_accounts_active_get_header_info())
    ws_mailinglist = wb.create_sheet(_("Mailing List"))
    ws_mailinglist.append(_export_excel_accounts_active_get_header_mailinglist())

    accounts = Account.objects.filter(is_active=True)

    for account in accounts:
        # Mailing List
        ws_mailinglist.append([
            account.first_name,
            account.last_name,
            account.email
        ])

        # Active accounts list
        ws_info.append([
            account.first_name,
            account.last_name,
            account.email,
            account.phone,
            account.mobile,
            account.emergency,
            account.date_of_birth,
            account.address,
            account.postcode,
            account.city,
            account.country,
            account.key_number
        ])

    # # Create a file-like buffer to receive xlsx data.
    buffer = io.BytesIO()
    wb.save(buffer)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)

    filename = f"Active_Accounts.xlsx"

    return FileResponse(buffer, as_attachment=True, filename=filename)
