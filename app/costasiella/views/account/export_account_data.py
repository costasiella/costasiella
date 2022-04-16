import io
import openpyxl

from django.utils.translation import gettext as _
from django.http import Http404, FileResponse

from ...models import AccountSubscriptionCredit, AccountInstructorProfile
# Django-graphql-jwt imports begin
from ...modules.graphql_jwt_tools import get_user_from_cookie
# Django-graphql-jwt imports end


def export_account_data(request, **kwargs):
    """
    Export account data

    :param: POST: token - JWT access token
    """
    user = get_user_from_cookie(request)
    if not user:
        raise Http404(_("Please sign in to export your account data."))

    # Create a new workbook to hold the worksheets containing data
    wb = openpyxl.workbook.Workbook(write_only=True)
    wb = _add_worksheet_account(user, wb)
    wb = _add_worksheet_account_accepted_documents(user, wb)
    wb = _add_worksheet_account_bank_account(user, wb)
    wb = _add_worksheet_account_bank_account_mandate(user, wb)
    wb = _add_worksheet_account_classpass(user, wb)
    wb = _add_worksheet_account_subscription(user, wb)
    wb = _add_worksheet_account_subscription_credit(user, wb)
    wb = _add_worksheet_account_event_tickets(user, wb)
    wb = _add_worksheet_account_note(user, wb)

    if user.instructor:
        wb = _add_worksheet_account_instructor_profile(user, wb)

    # # Create a file-like buffer to receive XLSX data.
    buffer = io.BytesIO()
    wb.save(buffer)

    buffer.seek(0)
    account_name = user.full_name.replace(" ", "_")
    filename = _('account_data') + '_' + account_name + ".xlsx"

    return FileResponse(buffer, as_attachment=True, filename=filename)


def _add_worksheet_account(account, wb):
    """
    Add Account sheet to workbook data export
    :param account: models.Account object
    :param wb: openpyxl.workbook.Workbook object
    :return: openpyxl.workbook.Workbook object (appended with Account worksheet)
    """
    ws = wb.create_sheet("Account")

    # Write header
    header = [
        "First name",
        "Last name",
        "Email",
        "Gender",
        "Date of birth",
        "Address",
        "Postcode",
        "City",
        "Country",
        "Phone",
        "Mobile",
        "Emergency",
        "Key nr",
        "Joined",
        "Last login",
    ]
    ws.append(header)

    # Write data
    data = [
        account.first_name,
        account.last_name,
        account.email,
        account.gender,
        account.date_of_birth,
        account.address,
        account.postcode,
        account.city,
        account.country,
        account.phone,
        account.mobile,
        account.emergency,
        account.key_number,
        account.date_joined.replace(tzinfo=None),
        account.last_login.replace(tzinfo=None) if account.last_login else "",
    ]
    ws.append(data)

    return wb


def _add_worksheet_account_instructor_profile(account, wb):
    """
    Add Account instructor profile sheet to workbook data export
    :param account: models.Account object
    :param wb: openpyxl.workbook.Workbook object
    :return: openpyxl.workbook.Workbook object (appended with Account worksheet)
    """
    ws = wb.create_sheet("Instructor profile")

    # Write header
    header = [
        "Teaches classes",
        "Teaches events",
        "Role",
        "Education",
        "Bio",
        "Bio URL",
        "Website URL"
    ]
    ws.append(header)

    qs = AccountInstructorProfile.objects.filter(account=account)
    if qs.exists():
        account_instructor_profile = qs.first()

        # Write data
        data = [
            account_instructor_profile.classes,
            account_instructor_profile.events,
            account_instructor_profile.role,
            account_instructor_profile.education,
            account_instructor_profile.bio,
            account_instructor_profile.url_bio,
            account_instructor_profile.url_website,
        ]
        ws.append(data)

    return wb


def _add_worksheet_account_accepted_documents(account, wb):
    """
    Add Account accepted documents sheet to workbook data export
    :param account: models.Account object
    :param wb: openpyxl.workbook.Workbook object
    :return: openpyxl.workbook.Workbook object (appended with Account worksheet)
    """
    ws = wb.create_sheet("Accepted documents")

    # Write header
    header = [
        "Document",
        "Version",
        "From IP",
        "On"
    ]
    ws.append(header)

    for accepted_document in account.accepted_documents.all():
        # Write data
        data = [
            accepted_document.document.document_type,
            accepted_document.document.version,
            accepted_document.client_ip,
            accepted_document.date_accepted.replace(tzinfo=None),
        ]
        ws.append(data)

    return wb


def _add_worksheet_account_bank_account(account, wb):
    """
    Add Account bank account sheet to workbook data export
    :param account: models.Account object
    :param wb: openpyxl.workbook.Workbook object
    :return: openpyxl.workbook.Workbook object (appended with Account worksheet)
    """
    ws = wb.create_sheet("Bank account")

    # Write header
    header = [
        "Accounr NR",
        "Holder",
        "BIC"
    ]
    ws.append(header)

    for bank_acount in account.bank_accounts.all():
        # Write data
        data = [
            bank_acount.number,
            bank_acount.holder,
            bank_acount.bic,
        ]
        ws.append(data)

    return wb


def _add_worksheet_account_bank_account_mandate(account, wb):
    """
    Add Account bank account mandate sheet to workbook data export
    :param account: models.Account object
    :param wb: openpyxl.workbook.Workbook object
    :return: openpyxl.workbook.Workbook object (appended with Account worksheet)
    """
    ws = wb.create_sheet("Mandates")

    # Write header
    header = [
        "Reference",
        "Content",
        "Sign date"
    ]
    ws.append(header)

    for bank_acount in account.bank_accounts.all():
        for mandate in bank_acount.mandates.all():
            # Write data
            data = [
                mandate.reference,
                mandate.content,
                mandate.signature_date
            ]
            ws.append(data)

    return wb


def _add_worksheet_account_classpass(account, wb):
    """
    Add Account classpass sheet to workbook data export
    :param account: models.Account object
    :param wb: openpyxl.workbook.Workbook object
    :return: openpyxl.workbook.Workbook object (appended with Account worksheet)
    """
    ws = wb.create_sheet("Classpasses")

    # Write header
    header = [
        "Classpass",
        "Start",
        "End",
        "Note",
        "Unlimited",
        "Classes remaining"
    ]
    ws.append(header)

    for classpass in account.classpasses.all():
        # Write data
        data = [
            classpass.organization_classpass.name,
            classpass.date_start,
            classpass.date_end,
            classpass.note,
            classpass.organization_classpass.unlimited,
            classpass.classes_remaining,
        ]
        ws.append(data)

    return wb


def _add_worksheet_account_subscription(account, wb):
    """
    Add Account subscription sheet to workbook data export
    :param account: models.Account object
    :param wb: openpyxl.workbook.Workbook object
    :return: openpyxl.workbook.Workbook object (appended with Account worksheet)
    """
    ws = wb.create_sheet("Subscriptions")

    # Write header
    header = [
        "Subscription #",
        "Subscription",
        "Start",
        "End",
        "Note",
    ]
    ws.append(header)

    for subscription in account.subscriptions.all():
        # Write data
        data = [
            subscription.id,
            subscription.organization_subscription.name,
            subscription.date_start,
            subscription.date_end,
            subscription.note,
        ]
        ws.append(data)

    return wb


def _add_worksheet_account_subscription_credit(account, wb):
    """
    Add Account subscription credits sheet to workbook data export
    :param account: models.Account object
    :param wb: openpyxl.workbook.Workbook object
    :return: openpyxl.workbook.Workbook object (appended with Account worksheet)
    """
    ws = wb.create_sheet("Subscription credits")

    # Write header
    header = [
        "Time",
        "Type",
        "Amount",
        "Description",
        "Subscription #",
    ]
    ws.append(header)

    qs = AccountSubscriptionCredit.objects.filter(
        account_subscription__account = account
    ).order_by("created_at")

    for subscription_credit in qs:
        # Write data
        data = [
            subscription_credit.created_at.replace(tzinfo=None),
            subscription_credit.mutation_type,
            subscription_credit.mutation_amount,
            subscription_credit.description,
            subscription_credit.account_subscription.id,
        ]
        ws.append(data)

    return wb


def _add_worksheet_account_event_tickets(account, wb):
    """
    Add Account event tickets sheet to workbook data export
    :param account: models.Account object
    :param wb: openpyxl.workbook.Workbook object
    :return: openpyxl.workbook.Workbook object (appended with Account worksheet)
    """
    ws = wb.create_sheet("Tickets")

    # Write header
    header = [
        "Event",
        "Ticket",
    ]
    ws.append(header)

    for ticket in account.schedule_event_tickets.all():
        # Write data
        data = [
            ticket.schedule_event_ticket.schedule_event.name,
            ticket.schedule_event_ticket.name,
        ]
        ws.append(data)

    return wb


def _add_worksheet_account_note(account, wb):
    """
    Add Account notes sheet to workbook data export
    :param account: models.Account object
    :param wb: openpyxl.workbook.Workbook object
    :return: openpyxl.workbook.Workbook object (appended with Account worksheet)
    """
    ws = wb.create_sheet("Notes")

    # Write header
    header = [
        "Time",
        "Note",
    ]
    ws.append(header)

    for note in account.notes.all():
        # Write data
        data = [
            note.created_at.replace(tzinfo=None),
            note.note
        ]
        ws.append(data)

    return wb
